import csv
import io
import uuid
from datetime import datetime, timedelta, timezone
from sqlalchemy import select, func, update, delete, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.timetable import (
    TimetableTerm, TimetableBatch, TimetableSection, TimetableLabGroup,
    TimetableRoom, TimetableCourse, TimetableFacultyProfile,
    TimetableStudentEnrollment, TimetableCourseOffering,
    TimetableTeacherEligibility, TimetableScheduleConfig,
    TimetableConstraint, TimetableSolveJob, TimetableEntry,
)
from app.models.routine import AcademicRoutine, RoutineStatus
from app.models.user import User
from app.schemas.timetable import (
    TermCreate, TermPatch,
    BatchCreate, BatchPatch, SectionCreate,
    RoomCreate, RoomPatch,
    CourseCreate, CoursePatch,
    FacultyProfileCreate, FacultyProfilePatch,
    OfferingCreate, EligibilityCreate,
    ScheduleConfigUpsert,
    ConstraintCreate, ConstraintPatch,
    EntryEdit, ConflictOut,
)

def _tenant_clause(col, tenant_id):
    """Same-tenant filter for duplicate checks; NULL tenant is its own scope."""
    return col == tenant_id if tenant_id is not None else col.is_(None)


# ── Terms ─────────────────────────────────────────────────────────────────────

async def list_terms(db: AsyncSession, tenant_id: uuid.UUID | None = None) -> list[TimetableTerm]:
    q = select(TimetableTerm)
    if tenant_id:
        q = q.where(TimetableTerm.tenant_id == tenant_id)
    result = await db.execute(q.order_by(TimetableTerm.created_at.desc()))
    return list(result.scalars().all())


async def create_term(db: AsyncSession, data: TermCreate) -> TimetableTerm:
    name = data.name.strip()
    dup = await db.execute(
        select(TimetableTerm)
        .where(_tenant_clause(TimetableTerm.tenant_id, data.tenant_id),
               func.lower(TimetableTerm.name) == name.lower())
        .limit(1)
    )
    if dup.scalars().first():
        # Duplicate term names made the term dropdown ambiguous and split
        # offerings across identical-looking terms (audit finding U-1/D-1).
        raise ValueError(f"A term named '{name}' already exists")
    term = TimetableTerm(name=name, tenant_id=data.tenant_id)
    db.add(term)
    await db.commit()
    await db.refresh(term)
    return term


async def get_term(db: AsyncSession, term_id: uuid.UUID) -> TimetableTerm | None:
    result = await db.execute(select(TimetableTerm).where(TimetableTerm.id == term_id))
    return result.scalars().first()


async def patch_term(db: AsyncSession, term: TimetableTerm, data: TermPatch) -> TimetableTerm:
    if data.name is not None:
        term.name = data.name
    if data.is_active is not None:
        if data.is_active:
            # deactivate all others in the same tenant
            all_terms = await list_terms(db, tenant_id=term.tenant_id)
            for t in all_terms:
                t.is_active = False
        term.is_active = data.is_active
    await db.commit()
    await db.refresh(term)
    return term


async def delete_term(db: AsyncSession, term: TimetableTerm) -> None:
    # guard: no solve jobs referencing this term
    result = await db.execute(select(TimetableSolveJob).where(TimetableSolveJob.term_id == term.id).limit(1))
    if result.scalars().first():
        raise ValueError("Cannot delete term with existing solve jobs")
    await db.delete(term)
    await db.commit()


# ── Batches / Sections / Lab groups ──────────────────────────────────────────

async def list_batches(db: AsyncSession, tenant_id: uuid.UUID | None = None) -> list[TimetableBatch]:
    q = select(TimetableBatch)
    if tenant_id:
        q = q.where(TimetableBatch.tenant_id == tenant_id)
    result = await db.execute(q.order_by(TimetableBatch.name))
    batches = list(result.scalars().all())
    # eagerly load sections + lab groups
    for batch in batches:
        sec_result = await db.execute(
            select(TimetableSection).where(TimetableSection.batch_id == batch.id)
        )
        batch._sections = list(sec_result.scalars().all())
        for sec in batch._sections:
            lg_result = await db.execute(
                select(TimetableLabGroup).where(TimetableLabGroup.section_id == sec.id)
            )
            sec._lab_groups = list(lg_result.scalars().all())
    return batches


async def create_batch(db: AsyncSession, data: BatchCreate) -> TimetableBatch:
    name = data.name.strip()
    rows = await db.execute(
        select(TimetableBatch.name).where(_tenant_clause(TimetableBatch.tenant_id, data.tenant_id))
    )
    if any(_norm_batch(n) == _norm_batch(name) for (n,) in rows.all()):
        raise ValueError(f"A batch named '{name}' already exists")
    batch = TimetableBatch(name=name, program=data.program, tenant_id=data.tenant_id)
    db.add(batch)
    await db.commit()
    await db.refresh(batch)
    return batch


async def get_batch(db: AsyncSession, batch_id: uuid.UUID) -> TimetableBatch | None:
    result = await db.execute(select(TimetableBatch).where(TimetableBatch.id == batch_id))
    return result.scalars().first()


async def patch_batch(db: AsyncSession, batch: TimetableBatch, data: BatchPatch) -> TimetableBatch:
    if data.name is not None:
        batch.name = data.name
    if data.program is not None:
        batch.program = data.program
    if data.active is not None:
        batch.active = data.active
    await db.commit()
    await db.refresh(batch)
    return batch


async def create_section(db: AsyncSession, batch_id: uuid.UUID, data: SectionCreate) -> TimetableSection:
    name = data.name.strip()
    dup = await db.execute(
        select(TimetableSection)
        .where(TimetableSection.batch_id == batch_id,
               func.lower(TimetableSection.name) == name.lower())
        .limit(1)
    )
    if dup.scalars().first():
        raise ValueError(f"Section '{name}' already exists in this batch")
    section = TimetableSection(batch_id=batch_id, name=name)
    db.add(section)
    await db.commit()
    await db.refresh(section)
    return section


async def create_lab_group(db: AsyncSession, section_id: uuid.UUID, name: str) -> TimetableLabGroup:
    lg = TimetableLabGroup(section_id=section_id, name=name)
    db.add(lg)
    await db.commit()
    await db.refresh(lg)
    return lg


async def generate_sections(
    db: AsyncSession, batch_id: uuid.UUID, count: int, lab_split: bool = True
) -> dict:
    """Idempotently ensure the batch has sections A..count (with lab groups).

    Only missing sections/lab groups are created — clicking "Generate" twice
    used to stack a second A, B, C… onto the batch (audit finding H-2).
    Returns {"created": [...names...], "existing": n}.
    """
    existing_r = await db.execute(
        select(TimetableSection).where(TimetableSection.batch_id == batch_id)
    )
    existing = {s.name: s for s in existing_r.scalars().all()}

    created: list[str] = []
    for i in range(count):
        name = chr(ord('A') + i)
        sec = existing.get(name)
        if sec is None:
            sec = TimetableSection(batch_id=batch_id, name=name)
            db.add(sec)
            await db.flush()
            created.append(name)
        if lab_split:
            lg_r = await db.execute(
                select(TimetableLabGroup.name).where(TimetableLabGroup.section_id == sec.id)
            )
            have = {n for (n,) in lg_r.all()}
            for lg_name in (f"{name}1", f"{name}2"):
                if lg_name not in have:
                    db.add(TimetableLabGroup(section_id=sec.id, name=lg_name))
    await db.commit()
    return {"created": created, "existing": len(existing)}


# ── Rooms ─────────────────────────────────────────────────────────────────────

async def list_rooms(db: AsyncSession, tenant_id: uuid.UUID | None = None) -> list[TimetableRoom]:
    q = select(TimetableRoom)
    if tenant_id:
        q = q.where(TimetableRoom.tenant_id == tenant_id)
    result = await db.execute(q.order_by(TimetableRoom.name))
    return list(result.scalars().all())


async def create_room(db: AsyncSession, data: RoomCreate) -> TimetableRoom:
    name = data.name.strip()
    dup = await db.execute(
        select(TimetableRoom)
        .where(_tenant_clause(TimetableRoom.tenant_id, data.tenant_id),
               func.lower(TimetableRoom.name) == name.lower())
        .limit(1)
    )
    if dup.scalars().first():
        raise ValueError(f"A room named '{name}' already exists")
    room = TimetableRoom(
        name=name, room_type=data.room_type,
        capacity=data.capacity, tenant_id=data.tenant_id,
    )
    db.add(room)
    await db.commit()
    await db.refresh(room)
    return room


async def get_room(db: AsyncSession, room_id: uuid.UUID) -> TimetableRoom | None:
    result = await db.execute(select(TimetableRoom).where(TimetableRoom.id == room_id))
    return result.scalars().first()


async def patch_room(db: AsyncSession, room: TimetableRoom, data: RoomPatch) -> TimetableRoom:
    if data.name is not None:
        room.name = data.name
    if data.room_type is not None:
        room.room_type = data.room_type
    if data.capacity is not None:
        room.capacity = data.capacity
    await db.commit()
    await db.refresh(room)
    return room


async def delete_room(db: AsyncSession, room: TimetableRoom) -> None:
    await db.delete(room)
    await db.commit()


# ── Courses ───────────────────────────────────────────────────────────────────

async def list_courses(db: AsyncSession, tenant_id: uuid.UUID | None = None) -> list[TimetableCourse]:
    q = select(TimetableCourse)
    if tenant_id:
        q = q.where(TimetableCourse.tenant_id == tenant_id)
    result = await db.execute(q.order_by(TimetableCourse.code))
    return list(result.scalars().all())


async def create_course(db: AsyncSession, data: CourseCreate) -> TimetableCourse:
    code = data.code.strip()
    rows = await db.execute(
        select(TimetableCourse.code).where(_tenant_clause(TimetableCourse.tenant_id, data.tenant_id))
    )
    if any(_norm_code(c) == _norm_code(code) for (c,) in rows.all()):
        raise ValueError(f"A course with code '{code}' already exists")
    course = TimetableCourse(
        code=code, name=data.name, credits=data.credits,
        is_lab=data.is_lab, weekly_classes=data.weekly_classes,
        tenant_id=data.tenant_id,
    )
    db.add(course)
    await db.commit()
    await db.refresh(course)
    return course


async def get_course(db: AsyncSession, course_id: uuid.UUID) -> TimetableCourse | None:
    result = await db.execute(select(TimetableCourse).where(TimetableCourse.id == course_id))
    return result.scalars().first()


async def patch_course(db: AsyncSession, course: TimetableCourse, data: CoursePatch) -> TimetableCourse:
    if data.code is not None:
        course.code = data.code
    if data.name is not None:
        course.name = data.name
    if data.credits is not None:
        course.credits = data.credits
    if data.is_lab is not None:
        course.is_lab = data.is_lab
    if data.weekly_classes is not None:
        course.weekly_classes = data.weekly_classes
    await db.commit()
    await db.refresh(course)
    return course


async def delete_course(db: AsyncSession, course: TimetableCourse) -> None:
    await db.delete(course)
    await db.commit()


# ── Faculty profiles ──────────────────────────────────────────────────────────

async def list_faculty(db: AsyncSession, tenant_id: uuid.UUID | None = None) -> list[TimetableFacultyProfile]:
    q = select(TimetableFacultyProfile)
    if tenant_id:
        q = q.where(TimetableFacultyProfile.tenant_id == tenant_id)
    result = await db.execute(q)
    return list(result.scalars().all())


async def create_faculty_profile(
    db: AsyncSession, data: FacultyProfileCreate
) -> TimetableFacultyProfile:
    name = (data.name or "").strip() or None
    email = (data.email or "").strip().lower() or None
    # When only a login account is supplied (manual "Add faculty" flow), derive
    # the display name / email from that user so faculty always have a label.
    if data.user_id and (not name or not email):
        u = (await db.execute(select(User).where(User.id == data.user_id))).scalars().first()
        if u:
            email = email or (u.email or "").strip().lower() or None
            name = name or (u.full_name or "").strip() or None
    name = name or (email.split("@")[0] if email else None)  # fall back to email prefix
    if email:
        dup = await db.execute(
            select(TimetableFacultyProfile)
            .where(_tenant_clause(TimetableFacultyProfile.tenant_id, data.tenant_id),
                   func.lower(TimetableFacultyProfile.email) == email)
            .limit(1)
        )
        if dup.scalars().first():
            raise ValueError(f"A faculty profile with email '{email}' already exists")
    fp = TimetableFacultyProfile(
        user_id=data.user_id, name=name, email=email, rank=data.rank,
        min_credits=data.min_credits, max_credits=data.max_credits,
        pref_slot=data.pref_slot, off_days=data.off_days,
        max_per_day=data.max_per_day, tenant_id=data.tenant_id,
    )
    db.add(fp)
    await db.commit()
    await db.refresh(fp)
    return fp


async def get_faculty_profile(db: AsyncSession, faculty_id: uuid.UUID) -> TimetableFacultyProfile | None:
    result = await db.execute(
        select(TimetableFacultyProfile).where(TimetableFacultyProfile.id == faculty_id)
    )
    return result.scalars().first()


async def patch_faculty_profile(
    db: AsyncSession, fp: TimetableFacultyProfile, data: FacultyProfilePatch
) -> TimetableFacultyProfile:
    if data.rank is not None:
        fp.rank = data.rank
    if data.min_credits is not None:
        fp.min_credits = data.min_credits
    if data.max_credits is not None:
        fp.max_credits = data.max_credits
    if data.pref_slot is not None:
        fp.pref_slot = data.pref_slot
    if data.off_days is not None:
        fp.off_days = data.off_days
    if data.max_per_day is not None:
        fp.max_per_day = data.max_per_day
    if data.active is not None:
        fp.active = data.active
    await db.commit()
    await db.refresh(fp)
    return fp


async def delete_faculty_profile(db: AsyncSession, fp: TimetableFacultyProfile) -> None:
    await db.delete(fp)
    await db.commit()


# ── Offerings ─────────────────────────────────────────────────────────────────

async def list_offerings(
    db: AsyncSession, term_id: uuid.UUID | None = None
) -> list[TimetableCourseOffering]:
    q = select(TimetableCourseOffering)
    if term_id:
        q = q.where(TimetableCourseOffering.term_id == term_id)
    offerings = list((await db.execute(q)).scalars().all())
    if offerings:
        # Attach human-readable labels so the admin table shows codes/names, not UUIDs.
        courses = {c.id: c.code for c in (await db.execute(select(TimetableCourse))).scalars().all()}
        batches = {b.id: b.name for b in (await db.execute(select(TimetableBatch))).scalars().all()}
        terms = {t.id: t.name for t in (await db.execute(select(TimetableTerm))).scalars().all()}
        for o in offerings:
            o.course_code = courses.get(o.course_id)
            o.batch_name = batches.get(o.batch_id)
            o.term_name = terms.get(o.term_id)
    return offerings


async def create_offering(db: AsyncSession, data: OfferingCreate) -> TimetableCourseOffering:
    dup = await db.execute(
        select(TimetableCourseOffering)
        .where(TimetableCourseOffering.term_id == data.term_id,
               TimetableCourseOffering.course_id == data.course_id,
               TimetableCourseOffering.batch_id == data.batch_id)
        .limit(1)
    )
    if dup.scalars().first():
        raise ValueError("This course is already offered to this batch in this term")
    offering = TimetableCourseOffering(
        term_id=data.term_id, course_id=data.course_id, batch_id=data.batch_id,
    )
    db.add(offering)
    await db.commit()
    await db.refresh(offering)
    return offering


async def delete_offering(db: AsyncSession, offering_id: uuid.UUID) -> None:
    result = await db.execute(
        select(TimetableCourseOffering).where(TimetableCourseOffering.id == offering_id)
    )
    obj = result.scalars().first()
    if obj:
        await db.delete(obj)
        await db.commit()


# ── Eligibility ───────────────────────────────────────────────────────────────

async def list_eligibility(
    db: AsyncSession, faculty_id: uuid.UUID | None = None, course_id: uuid.UUID | None = None
) -> list[TimetableTeacherEligibility]:
    q = select(TimetableTeacherEligibility)
    if faculty_id:
        q = q.where(TimetableTeacherEligibility.faculty_id == faculty_id)
    if course_id:
        q = q.where(TimetableTeacherEligibility.course_id == course_id)
    eligs = list((await db.execute(q)).scalars().all())
    if eligs:
        # Attach human-readable labels for the admin table.
        courses = {c.id: c.code for c in (await db.execute(select(TimetableCourse))).scalars().all()}
        faculty = {f.id: f for f in (await db.execute(select(TimetableFacultyProfile))).scalars().all()}
        for e in eligs:
            e.course_code = courses.get(e.course_id)
            fp = faculty.get(e.faculty_id)
            e.faculty_name = (fp.name or fp.email) if fp else None
    return eligs


async def create_eligibility(
    db: AsyncSession, data: EligibilityCreate
) -> TimetableTeacherEligibility:
    dup = await db.execute(
        select(TimetableTeacherEligibility)
        .where(TimetableTeacherEligibility.faculty_id == data.faculty_id,
               TimetableTeacherEligibility.course_id == data.course_id)
        .limit(1)
    )
    if dup.scalars().first():
        raise ValueError("This teacher is already eligible for this course")
    elig = TimetableTeacherEligibility(faculty_id=data.faculty_id, course_id=data.course_id)
    db.add(elig)
    await db.commit()
    await db.refresh(elig)
    return elig


async def delete_eligibility(db: AsyncSession, elig_id: uuid.UUID) -> None:
    result = await db.execute(
        select(TimetableTeacherEligibility).where(TimetableTeacherEligibility.id == elig_id)
    )
    obj = result.scalars().first()
    if obj:
        await db.delete(obj)
        await db.commit()


# ── Schedule config ───────────────────────────────────────────────────────────

async def get_schedule_config(
    db: AsyncSession, term_id: uuid.UUID | None = None, tenant_id: uuid.UUID | None = None
) -> TimetableScheduleConfig | None:
    q = select(TimetableScheduleConfig)
    if term_id:
        q = q.where(TimetableScheduleConfig.term_id == term_id)
    elif tenant_id:
        q = q.where(TimetableScheduleConfig.tenant_id == tenant_id)
    result = await db.execute(q.limit(1))
    return result.scalars().first()


async def upsert_schedule_config(
    db: AsyncSession, data: ScheduleConfigUpsert, requesting_tenant: uuid.UUID | None = None
) -> TimetableScheduleConfig:
    existing = await get_schedule_config(db, term_id=data.term_id, tenant_id=requesting_tenant)
    if existing:
        existing.days = data.days
        existing.slots = data.slots
        existing.off_days = data.off_days
        await db.commit()
        await db.refresh(existing)
        return existing
    cfg = TimetableScheduleConfig(
        tenant_id=data.tenant_id or requesting_tenant,
        term_id=data.term_id,
        days=data.days,
        slots=data.slots,
        off_days=data.off_days,
    )
    db.add(cfg)
    await db.commit()
    await db.refresh(cfg)
    return cfg


# ── Constraints ───────────────────────────────────────────────────────────────

async def list_constraints(
    db: AsyncSession,
    term_id: uuid.UUID | None = None,
    tenant_id: uuid.UUID | None = None,
) -> list[TimetableConstraint]:
    q = select(TimetableConstraint)
    if term_id:
        q = q.where(TimetableConstraint.term_id == term_id)
    if tenant_id:
        q = q.where(TimetableConstraint.tenant_id == tenant_id)
    result = await db.execute(q.order_by(TimetableConstraint.enforcement, TimetableConstraint.constraint_type))
    return list(result.scalars().all())


async def create_constraint(
    db: AsyncSession, data: ConstraintCreate, created_by: uuid.UUID | None = None
) -> TimetableConstraint:
    con = TimetableConstraint(
        constraint_type=data.constraint_type,
        scope=data.scope, params=data.params,
        enforcement=data.enforcement, weight=data.weight,
        enabled=True,
        term_id=data.term_id, tenant_id=data.tenant_id,
        created_by=created_by,
    )
    db.add(con)
    await db.commit()
    await db.refresh(con)
    return con


async def get_constraint(db: AsyncSession, con_id: uuid.UUID) -> TimetableConstraint | None:
    result = await db.execute(select(TimetableConstraint).where(TimetableConstraint.id == con_id))
    return result.scalars().first()


async def patch_constraint(
    db: AsyncSession, con: TimetableConstraint, data: ConstraintPatch
) -> TimetableConstraint:
    if data.enabled is not None:
        con.enabled = data.enabled
    if data.weight is not None:
        con.weight = data.weight
    if data.enforcement is not None:
        con.enforcement = data.enforcement
    if data.params is not None:
        con.params = data.params
    await db.commit()
    await db.refresh(con)
    return con


async def delete_constraint(db: AsyncSession, con: TimetableConstraint) -> None:
    await db.delete(con)
    await db.commit()


# ── Solve jobs ────────────────────────────────────────────────────────────────

async def create_solve_job(
    db: AsyncSession, term_id: uuid.UUID,
    requested_by: uuid.UUID | None = None,
    tenant_id: uuid.UUID | None = None,
    params: dict | None = None,
) -> TimetableSolveJob:
    job = TimetableSolveJob(
        term_id=term_id, requested_by=requested_by,
        tenant_id=tenant_id, params=params or {},
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)
    return job


async def get_solve_job(db: AsyncSession, job_id: uuid.UUID) -> TimetableSolveJob | None:
    result = await db.execute(select(TimetableSolveJob).where(TimetableSolveJob.id == job_id))
    return result.scalars().first()


async def update_solve_job(db: AsyncSession, job: TimetableSolveJob, **kwargs) -> TimetableSolveJob:
    for k, v in kwargs.items():
        setattr(job, k, v)
    await db.commit()
    await db.refresh(job)
    return job


async def reap_stale_solve_jobs(db: AsyncSession, older_than_s: int = 60) -> int:
    """Mark queued/running jobs with no recent heartbeat as orphaned.

    Called at startup: any job still 'running' from before a restart is dead —
    its background task did not survive the process. `updated_at` doubles as
    the heartbeat because the solver writes progress to the row per batch
    group. The small grace window protects a job another worker just started.
    """
    cutoff = datetime.now(timezone.utc) - timedelta(seconds=older_than_s)
    result = await db.execute(
        update(TimetableSolveJob)
        .where(
            TimetableSolveJob.status.in_(("queued", "running")),
            TimetableSolveJob.updated_at < cutoff,
        )
        .values(
            status="failed",
            solver_status="orphaned",
            progress=100,
            finished_at=datetime.now(timezone.utc),
        )
        # Row-level operation: don't let the ORM "evaluate" the WHERE against
        # in-session objects (its naive-vs-aware datetime compare misfires).
        .execution_options(synchronize_session=False)
    )
    await db.commit()
    return result.rowcount or 0


# ── Entries ───────────────────────────────────────────────────────────────────

async def list_entries(
    db: AsyncSession, term_id: uuid.UUID, version: int
) -> list[TimetableEntry]:
    result = await db.execute(
        select(TimetableEntry)
        .where(TimetableEntry.term_id == term_id, TimetableEntry.result_version == version)
        .order_by(TimetableEntry.day, TimetableEntry.slot)
    )
    return list(result.scalars().all())


async def get_entry(db: AsyncSession, entry_id: uuid.UUID) -> TimetableEntry | None:
    result = await db.execute(select(TimetableEntry).where(TimetableEntry.id == entry_id))
    return result.scalars().first()


async def patch_entry(
    db: AsyncSession,
    entry: TimetableEntry,
    data: EntryEdit,
    config: TimetableScheduleConfig | None = None,
) -> TimetableEntry:
    if config is not None:
        if data.new_day is not None and not (0 <= data.new_day < len(config.days)):
            raise ValueError(
                f"day index {data.new_day} is out of range — this term has {len(config.days)} days"
            )
        if data.new_slot is not None and not (0 <= data.new_slot < len(config.slots)):
            raise ValueError(
                f"slot index {data.new_slot} is out of range — this term has {len(config.slots)} slots"
            )
    if data.new_day is not None:
        entry.day = data.new_day
    if data.new_slot is not None:
        entry.slot = data.new_slot
    if data.new_faculty_id is not None:
        entry.faculty_id = data.new_faculty_id
    if data.new_room_id is not None:
        entry.room_id = data.new_room_id
    entry.locked = data.lock
    entry.source = "manual"
    await db.commit()
    await db.refresh(entry)
    return entry


async def get_next_version(db: AsyncSession, term_id: uuid.UUID) -> int:
    result = await db.execute(
        select(func.max(TimetableEntry.result_version))
        .where(TimetableEntry.term_id == term_id)
    )
    current_max = result.scalar()
    return (current_max or 0) + 1


async def bulk_insert_entries(
    db: AsyncSession, entries: list[TimetableEntry]
) -> None:
    for entry in entries:
        db.add(entry)
    await db.commit()


# ── Bulk name resolution ──────────────────────────────────────────────────────

async def resolve_entry_names(db: AsyncSession, entries: list[TimetableEntry]) -> dict:
    """Bulk-load the entities a set of entries references (one IN query per
    table instead of one query per entry). Returns lookup maps keyed by id:
    courses, sections, batches, lab_groups, rooms, faculty, plus
    faculty_names: {faculty_profile_id: user full name}."""

    async def _by_ids(model, ids):
        ids = [i for i in ids if i]
        if not ids:
            return {}
        r = await db.execute(select(model).where(model.id.in_(ids)))
        return {obj.id: obj for obj in r.scalars().all()}

    courses = await _by_ids(TimetableCourse, {e.course_id for e in entries})
    sections = await _by_ids(TimetableSection, {e.section_id for e in entries})
    batches = await _by_ids(TimetableBatch, {s.batch_id for s in sections.values()})
    lab_groups = await _by_ids(TimetableLabGroup, {e.lab_group_id for e in entries if e.lab_group_id})
    rooms = await _by_ids(TimetableRoom, {e.room_id for e in entries})
    faculty = await _by_ids(TimetableFacultyProfile, {e.faculty_id for e in entries})

    users = {}
    user_ids = [fp.user_id for fp in faculty.values() if fp.user_id]
    if user_ids:
        ur = await db.execute(select(User).where(User.id.in_(user_ids)))
        users = {u.id: u for u in ur.scalars().all()}

    def _fac_name(fp):
        # Prefer the faculty's own name; fall back to a linked account, then the
        # email prefix, so account-less imported faculty still show a label.
        if fp.name:
            return fp.name
        if fp.user_id in users:
            return users[fp.user_id].full_name
        return (fp.email.split("@")[0] if fp.email else "")

    faculty_names = {fid: _fac_name(fp) for fid, fp in faculty.items()}

    return {
        "courses": courses,
        "sections": sections,
        "batches": batches,
        "lab_groups": lab_groups,
        "rooms": rooms,
        "faculty": faculty,
        "faculty_names": faculty_names,
    }


# ── Validate entries ──────────────────────────────────────────────────────────

async def validate_entries(
    db: AsyncSession, term_id: uuid.UUID, version: int
) -> list[ConflictOut]:
    entries = await list_entries(db, term_id, version)
    conflicts: list[ConflictOut] = []

    # Load room types (single IN query)
    room_ids = [rid for rid in {e.room_id for e in entries} if rid]
    rooms: dict = {}
    if room_ids:
        r = await db.execute(select(TimetableRoom).where(TimetableRoom.id.in_(room_ids)))
        rooms = {room.id: room for room in r.scalars().all()}

    # Check (room, day, slot) uniqueness
    room_slot_seen: dict = {}
    for e in entries:
        key = (e.room_id, e.day, e.slot)
        if key in room_slot_seen:
            prev = room_slot_seen[key]
            conflicts.append(ConflictOut(
                conflict_type="room_double_booked",
                entry_ids=[prev.id, e.id],
                description=f"Room {rooms.get(e.room_id, {}).name if e.room_id in rooms else e.room_id} double-booked on day {e.day} slot {e.slot}",
            ))
        else:
            room_slot_seen[key] = e

    # Check (faculty, day, slot) uniqueness
    faculty_slot_seen: dict = {}
    for e in entries:
        key = (e.faculty_id, e.day, e.slot)
        if key in faculty_slot_seen:
            prev = faculty_slot_seen[key]
            conflicts.append(ConflictOut(
                conflict_type="teacher_overlap",
                entry_ids=[prev.id, e.id],
                description=f"Teacher assigned to two classes on day {e.day} slot {e.slot}",
            ))
        else:
            faculty_slot_seen[key] = e

    # Check theory section (day, slot) uniqueness — labs use lab_group so they're separate
    theory_slot_seen: dict = {}
    for e in entries:
        if not e.is_lab:
            key = (e.section_id, e.day, e.slot)
            if key in theory_slot_seen:
                prev = theory_slot_seen[key]
                conflicts.append(ConflictOut(
                    conflict_type="section_overlap",
                    entry_ids=[prev.id, e.id],
                    description=f"Section has two theory classes on day {e.day} slot {e.slot}",
                ))
            else:
                theory_slot_seen[key] = e

    # Check the same lab group isn't double-booked — the solver prevents this,
    # but manual entry edits can create it and nothing else would catch it.
    lab_group_slot_seen: dict = {}
    for e in entries:
        if e.is_lab and e.lab_group_id is not None:
            key = (e.lab_group_id, e.day, e.slot)
            if key in lab_group_slot_seen:
                prev = lab_group_slot_seen[key]
                conflicts.append(ConflictOut(
                    conflict_type="lab_group_overlap",
                    entry_ids=[prev.id, e.id],
                    description=f"Lab group has two classes on day {e.day} slot {e.slot}",
                ))
            else:
                lab_group_slot_seen[key] = e

    # Check room type matches is_lab (labs need a LAB room; theory must not
    # occupy one — THEORY or ONLINE are both fine for theory)
    for e in entries:
        room = rooms.get(e.room_id)
        if room:
            if e.is_lab and room.room_type != "LAB":
                conflicts.append(ConflictOut(
                    conflict_type="room_type_mismatch",
                    entry_ids=[e.id],
                    description=f"Lab class assigned to {room.room_type} room {room.name}",
                ))
            elif not e.is_lab and room.room_type == "LAB":
                conflicts.append(ConflictOut(
                    conflict_type="room_type_mismatch",
                    entry_ids=[e.id],
                    description=f"Theory class assigned to LAB room {room.name}",
                ))

    # Check a section's theory and its own lab don't share a slot — lab-group
    # students also attend the section theory, so it's a real student clash.
    # (Two different lab groups of the same section may coexist.)
    section_theory_slot: dict = {}
    for e in entries:
        if not e.is_lab:
            section_theory_slot[(e.section_id, e.day, e.slot)] = e
    for e in entries:
        if e.is_lab:
            theory = section_theory_slot.get((e.section_id, e.day, e.slot))
            if theory is not None:
                conflicts.append(ConflictOut(
                    conflict_type="section_lab_overlap",
                    entry_ids=[theory.id, e.id],
                    description=f"Section has a theory and a lab class at the same time on day {e.day} slot {e.slot}",
                ))

    # Check lab entries have a lab_group_id
    for e in entries:
        if e.is_lab and e.lab_group_id is None:
            conflicts.append(ConflictOut(
                conflict_type="missing_lab_group",
                entry_ids=[e.id],
                description="Lab entry missing lab_group_id",
            ))

    return conflicts


# ── Publish entries → AcademicRoutine ─────────────────────────────────────────

async def publish_version(
    db: AsyncSession,
    term_id: uuid.UUID,
    version: int,
    published_by: uuid.UUID,
) -> list[uuid.UUID]:
    """Project tt_entries → AcademicRoutine rows (one per section), publish them."""
    entries = await list_entries(db, term_id, version)
    if not entries:
        raise ValueError("No entries found for this term/version")

    # Load term
    term_result = await db.execute(select(TimetableTerm).where(TimetableTerm.id == term_id))
    term = term_result.scalars().first()
    if not term:
        raise ValueError("Term not found")

    # Load schedule config for day/slot name resolution
    config = await get_schedule_config(db, term_id=term_id)
    if not config:
        raise ValueError("Schedule config not found for this term — set days/slots first")

    days_list: list[str] = config.days
    slots_list: list[str] = config.slots

    # Load supporting data in bulk (batch via section, course, faculty→user, room)
    names = await resolve_entry_names(db, entries)
    sections = names["sections"]
    batches = names["batches"]
    courses = names["courses"]
    faculty_map = names["faculty_names"]
    room_map = {rid: r.name for rid, r in names["rooms"].items()}
    lab_group_map = {lgid: lg.name for lgid, lg in names["lab_groups"].items()}

    # Group entries by section
    by_section: dict[uuid.UUID, list[TimetableEntry]] = {}
    for e in entries:
        by_section.setdefault(e.section_id, []).append(e)

    published_ids: list[uuid.UUID] = []
    now = datetime.now(timezone.utc)

    for section_id, sec_entries in by_section.items():
        section = sections.get(section_id)
        if not section:
            continue
        batch = batches.get(section.batch_id)
        if not batch:
            continue

        # Build SlotItem list (matches app/schemas/routines.py SlotItem shape)
        slot_items = []
        for e in sec_entries:
            day_name = days_list[e.day] if e.day < len(days_list) else str(e.day)
            slot_str = slots_list[e.slot] if e.slot < len(slots_list) else str(e.slot)
            # Parse "8:30-10:00" into start/end
            if "-" in slot_str:
                parts = slot_str.split("-", 1)
                start_time, end_time = parts[0].strip(), parts[1].strip()
            else:
                start_time, end_time = slot_str, ""

            teacher_name = faculty_map.get(e.faculty_id, "")
            room_name = room_map.get(e.room_id, "")
            lab_suffix = f" [{lab_group_map.get(e.lab_group_id, '')}]" if e.lab_group_id else ""

            course = courses.get(e.course_id)

            slot_items.append({
                "day": day_name,
                "start_time": start_time,
                "end_time": end_time,
                "course_code": course.code if course else "",
                "course_name": (course.name if course else "") + lab_suffix,
                "teacher": teacher_name,
                "room": room_name,
            })

        dept = batch.program
        # Encode section into batch so A/B/C each get their own AcademicRoutine row
        batch_name = f"{batch.name} ({section.name})"
        semester = term.name
        title = f"{dept} {section.name} — {term.name}"

        # Upsert: find existing non-archived routine for this section identity
        existing_q = (
            select(AcademicRoutine)
            .where(
                AcademicRoutine.department == dept,
                AcademicRoutine.batch == batch_name,
                AcademicRoutine.semester == semester,
                AcademicRoutine.status != RoutineStatus.archived,
            )
        )
        existing_r = await db.execute(existing_q)
        routine = existing_r.scalars().first()

        if routine is None:
            routine = AcademicRoutine(
                tenant_id=term.tenant_id,
                department=dept,
                batch=batch_name,
                semester=semester,
                academic_year="",
                title=title,
                slots=slot_items,
                created_by_user_id=published_by,
            )
            db.add(routine)
        else:
            routine.slots = slot_items
            routine.title = title

        routine.status = RoutineStatus.published
        routine.published_by_user_id = published_by
        routine.published_at = now
        await db.flush()
        published_ids.append(routine.id)

    # Mark entries as published
    for e in entries:
        e.published = True

    await db.commit()
    return published_ids


# ── CSV / XLSX import ─────────────────────────────────────────────────────────

def _int_or_none(val) -> int | None:
    val = str(val or "").strip()
    return int(val) if val else None


# Minimum columns each entity import needs. Missing any of these is reported once
# (a clear header error) instead of once per data row.
_REQUIRED_COLS: dict[str, tuple[str, ...]] = {
    "courses": ("code", "name"),
    "rooms": ("name",),
    "batches": ("name",),
    "faculty": ("email",),
    "offerings": ("course_code", "batch_name"),
    "eligibility": ("faculty_email", "course_code"),
}

# Full expected header, shown in the "missing column" message to guide the admin.
_EXPECTED_HEADER: dict[str, str] = {
    "courses": "code, name, credits, weekly_classes, is_lab",
    "rooms": "name, room_type, capacity",
    "batches": "name, program",
    "faculty": "email, rank, max_per_day",
    "offerings": "course_code, batch_name",
    "eligibility": "faculty_email, course_code",
}


def _norm_code(s) -> str:
    """Normalise a course code for forgiving matching: case-, space-, hyphen-insensitive.

    So 'SE 111', 'se-111', and 'SE111' all resolve to the same course.
    """
    return "".join(str(s or "").split()).replace("-", "").replace("_", "").lower()


def _norm_batch(s) -> str:
    """Normalise a batch name for forgiving matching.

    Tolerates a leading 'batch' word and any spacing/hyphens, so an offerings
    file that says '38', 'Batch 38', 'Batch-38', or 'batch38' all resolve to
    the batch stored as 'Batch 38'.
    """
    x = "".join(str(s or "").split()).replace("-", "").replace("_", "").lower()
    if x.startswith("batch"):
        x = x[len("batch"):]
    return x


def _sniff_delimiter(text: str) -> str:
    """Detect the CSV delimiter (comma, semicolon, tab, or pipe).

    Excel in some locales exports CSV with ';'; without this every row lands in
    a single column and the whole import fails. Falls back to the most common
    candidate on the header line, then to comma.
    """
    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        header = sample.splitlines()[0] if sample.splitlines() else ""
        counts = {d: header.count(d) for d in (",", ";", "\t", "|")}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


async def import_entities(
    db: AsyncSession,
    entity: str,
    file_bytes: bytes,
    content_type: str,
    tenant_id: uuid.UUID | None = None,
    term_id: uuid.UUID | None = None,
) -> dict:
    fieldnames, rows = _parse_file(file_bytes, content_type)
    created = 0
    error_count = 0
    errors: list[str] = []
    MAX_ERRORS = 50  # cap the returned list so a huge bad file stays lightweight

    required = _REQUIRED_COLS.get(entity)
    if required is None:
        return {"created": 0, "total": 0, "error_count": 1,
                "errors": [f"unsupported entity '{entity}'"]}

    # Header validation up front: collapse "every row fails with a cryptic
    # KeyError" into one actionable message naming the missing/expected columns.
    field_set = {f for f in fieldnames if f}
    missing = [c for c in required if c not in field_set]
    if missing:
        found = ", ".join(f for f in fieldnames if f) or "(none)"
        return {
            "created": 0,
            "total": len(rows),
            "error_count": 1,
            "errors": [
                f"Missing required column(s): {', '.join(missing)}. "
                f"Found: {found}. Expected header: {_EXPECTED_HEADER.get(entity, '')}"
            ],
        }

    # Pre-load reference lookups (human-readable keys → UUIDs) so admins never
    # have to paste IDs into a spreadsheet, and so re-importing a file UPDATES
    # matching rows instead of appending duplicates (audit finding H-1 — this
    # append behavior is what doubled the production course table). We store
    # scalar IDs (not ORM objects) because the per-row rollback below expires
    # all instances in the session — holding IDs keeps the lookups valid.
    course_id_by_code: dict[str, uuid.UUID] = {}
    batch_id_by_name: dict[str, uuid.UUID] = {}
    room_id_by_name: dict[str, uuid.UUID] = {}
    user_id_by_email: dict[str, uuid.UUID] = {}
    faculty_id_by_email: dict[str, uuid.UUID] = {}
    existing_offerings: set[tuple] = set()
    existing_eligibility: set[tuple] = set()
    course_codes_display: list[str] = []   # original codes, for error messages
    batch_names_display: list[str] = []     # original names, for error messages

    if entity in ("courses", "offerings", "eligibility"):
        _courses = await list_courses(db, tenant_id)
        course_id_by_code = {_norm_code(c.code): c.id for c in _courses}
        course_codes_display = [c.code for c in _courses]
    if entity in ("batches", "offerings"):
        _batches = await list_batches(db, tenant_id)
        batch_id_by_name = {_norm_batch(b.name): b.id for b in _batches}
        batch_names_display = [b.name for b in _batches]
    if entity == "rooms":
        room_id_by_name = {r.name.strip().lower(): r.id for r in await list_rooms(db, tenant_id)}
    if entity in ("faculty", "eligibility"):
        result = await db.execute(select(User))
        user_id_by_email = {u.email.strip().lower(): u.id for u in result.scalars().all() if u.email}
    if entity in ("faculty", "eligibility"):
        # Resolve faculty by their own email, falling back to a linked account's
        # email so eligibility CSVs work whether or not the teacher has a login.
        email_by_user_id = {uid: em for em, uid in user_id_by_email.items()}
        for fp in await list_faculty(db, tenant_id):
            key = (fp.email or "").strip().lower() or email_by_user_id.get(fp.user_id)
            if key:
                faculty_id_by_email[key] = fp.id
    if entity == "offerings" and term_id:
        result = await db.execute(
            select(TimetableCourseOffering).where(TimetableCourseOffering.term_id == term_id)
        )
        existing_offerings = {(o.course_id, o.batch_id) for o in result.scalars().all()}
    if entity == "eligibility":
        result = await db.execute(select(TimetableTeacherEligibility))
        existing_eligibility = {(e.faculty_id, e.course_id) for e in result.scalars().all()}

    updated = 0
    skipped = 0

    for i, row in enumerate(rows, start=2):  # row 1 = header
        try:
            outcome = "created"
            # Keys are already stripped + lowercased by _parse_file.
            if entity == "courses":
                code = (row.get("code") or "").strip()
                name = (row.get("name") or "").strip()
                if not code or not name:
                    raise ValueError("missing value for 'code' or 'name'")
                values = dict(
                    name=name,
                    credits=int(row.get("credits") or 3),
                    is_lab=str(row.get("is_lab") or "false").strip().lower() in ("true", "1", "yes", "y"),
                    weekly_classes=int(row.get("weekly_classes") or 2),
                )
                existing_id = course_id_by_code.get(_norm_code(code))
                if existing_id is not None:
                    await db.execute(
                        update(TimetableCourse).where(TimetableCourse.id == existing_id).values(**values)
                    )
                    await db.commit()
                    outcome = "updated"
                else:
                    c = await create_course(db, CourseCreate(code=code, tenant_id=tenant_id, **values))
                    course_id_by_code[_norm_code(code)] = c.id
            elif entity == "rooms":
                name = (row.get("name") or "").strip()
                if not name:
                    raise ValueError("missing value for 'name'")
                values = dict(
                    room_type=((row.get("room_type") or "THEORY").strip().upper() or "THEORY"),
                    capacity=int(row.get("capacity") or 30),
                )
                existing_id = room_id_by_name.get(name.lower())
                if existing_id is not None:
                    await db.execute(
                        update(TimetableRoom).where(TimetableRoom.id == existing_id).values(**values)
                    )
                    await db.commit()
                    outcome = "updated"
                else:
                    r = await create_room(db, RoomCreate(name=name, tenant_id=tenant_id, **values))
                    room_id_by_name[name.lower()] = r.id
            elif entity == "batches":
                name = (row.get("name") or "").strip()
                if not name:
                    raise ValueError("missing value for 'name'")
                program = (row.get("program") or "SWE").strip() or "SWE"
                existing_id = batch_id_by_name.get(_norm_batch(name))
                if existing_id is not None:
                    await db.execute(
                        update(TimetableBatch).where(TimetableBatch.id == existing_id).values(program=program)
                    )
                    await db.commit()
                    outcome = "updated"
                else:
                    b = await create_batch(db, BatchCreate(name=name, program=program, tenant_id=tenant_id))
                    batch_id_by_name[_norm_batch(name)] = b.id
            elif entity == "faculty":
                email = str(row.get("email") or "").strip().lower()
                if not email:
                    raise ValueError("missing value for 'email'")
                values = dict(
                    rank=str(row.get("rank") or "").strip().upper() or "LECTURER",
                    max_per_day=int(row.get("max_per_day") or 4),
                    min_credits=_int_or_none(row.get("min_credits")),
                    max_credits=_int_or_none(row.get("max_credits")),
                    pref_slot=_int_or_none(row.get("pref_slot")),
                )
                name = str(row.get("name") or "").strip() or None
                existing_id = faculty_id_by_email.get(email)
                if existing_id is not None:
                    if name:
                        values["name"] = name
                    await db.execute(
                        update(TimetableFacultyProfile)
                        .where(TimetableFacultyProfile.id == existing_id)
                        .values(**values)
                    )
                    await db.commit()
                    outcome = "updated"
                else:
                    fp = await create_faculty_profile(db, FacultyProfileCreate(
                        user_id=user_id_by_email.get(email),  # link a login account if one exists, else None
                        name=name, email=email, tenant_id=tenant_id, **values,
                    ))
                    faculty_id_by_email[email] = fp.id
            elif entity == "offerings":
                if not term_id:
                    raise ValueError("no term selected — pick a term before importing offerings")
                raw_code = str(row.get("course_code") or "").strip()
                course_id = course_id_by_code.get(_norm_code(raw_code))
                if course_id is None:
                    avail = ", ".join(sorted(course_codes_display)[:8]) or "(none)"
                    raise ValueError(
                        f"no course with code '{raw_code}' — {len(course_id_by_code)} course(s) "
                        f"visible to this account: {avail}. Import Courses first (under the same login)."
                    )
                raw_batch = str(row.get("batch_name") or "").strip()
                batch_id = batch_id_by_name.get(_norm_batch(raw_batch))
                if batch_id is None:
                    avail = ", ".join(sorted(batch_names_display)[:8]) or "(none)"
                    raise ValueError(
                        f"no batch named '{raw_batch}' — {len(batch_id_by_name)} batch(es) "
                        f"visible to this account: {avail}. Import Batches first (under the same login)."
                    )
                if (course_id, batch_id) in existing_offerings:
                    outcome = "skipped"
                else:
                    await create_offering(db, OfferingCreate(
                        term_id=term_id, course_id=course_id, batch_id=batch_id,
                    ))
                    existing_offerings.add((course_id, batch_id))
            elif entity == "eligibility":
                email = str(row.get("faculty_email") or "").strip().lower()
                if not email:
                    raise ValueError("missing value for 'faculty_email'")
                fp_id = faculty_id_by_email.get(email)
                if fp_id is None:
                    avail = ", ".join(sorted(faculty_id_by_email)[:5]) or "(none)"
                    raise ValueError(
                        f"no faculty found with email '{email}' — {len(faculty_id_by_email)} "
                        f"faculty visible to this account: {avail}. Import Faculty first (under the same login)."
                    )
                raw_code = str(row.get("course_code") or "").strip()
                course_id = course_id_by_code.get(_norm_code(raw_code))
                if course_id is None:
                    avail = ", ".join(sorted(course_codes_display)[:8]) or "(none)"
                    raise ValueError(
                        f"no course with code '{raw_code}' — {len(course_id_by_code)} course(s) "
                        f"visible to this account: {avail}. Import Courses first (under the same login)."
                    )
                if (fp_id, course_id) in existing_eligibility:
                    outcome = "skipped"
                else:
                    await create_eligibility(db, EligibilityCreate(
                        faculty_id=fp_id, course_id=course_id,
                    ))
                    existing_eligibility.add((fp_id, course_id))

            if outcome == "created":
                created += 1
            elif outcome == "updated":
                updated += 1
            else:
                skipped += 1
        except Exception as exc:
            try:
                await db.rollback()  # clear any failed txn so one bad row can't cascade
            except Exception:
                pass
            error_count += 1
            if len(errors) < MAX_ERRORS:
                errors.append(f"Row {i}: {exc}")

    return {
        "created": created, "updated": updated, "skipped": skipped,
        "total": len(rows), "error_count": error_count, "errors": errors,
    }


async def clear_entities(
    db: AsyncSession,
    entity: str,
    tenant_id: uuid.UUID | None = None,
    term_id: uuid.UUID | None = None,
) -> dict[str, int]:
    """Bulk-delete every row of ``entity`` visible to this admin (the "Remove
    import" button). Dependent rows are deleted first in FK order — the tables
    have no ON DELETE CASCADE, and leaving e.g. timetable entries pointing at a
    deleted course would 500 every subsequent solve/list anyway.

    Scoping mirrors the list_* helpers: a tenant admin only clears rows with
    their tenant_id; a tenant-less account clears everything it can see.
    Offerings additionally scope to ``term_id`` when given (imports are
    per-term, so removal is too).

    Returns per-table deleted counts, e.g. {"courses": 12, "tt_entries": 96}.
    """
    deleted: dict[str, int] = {}

    async def _del(label: str, stmt) -> None:
        res = await db.execute(stmt)
        n = res.rowcount or 0
        if n:
            deleted[label] = deleted.get(label, 0) + n

    def _scoped_ids(model):
        q = select(model.id)
        if tenant_id:
            q = q.where(model.tenant_id == tenant_id)
        return q

    course_ids = _scoped_ids(TimetableCourse)
    batch_ids = _scoped_ids(TimetableBatch)
    room_ids = _scoped_ids(TimetableRoom)
    faculty_ids = _scoped_ids(TimetableFacultyProfile)
    term_ids = _scoped_ids(TimetableTerm)
    section_ids = select(TimetableSection.id).where(TimetableSection.batch_id.in_(batch_ids))

    if entity == "courses":
        await _del("eligibility", delete(TimetableTeacherEligibility)
                   .where(TimetableTeacherEligibility.course_id.in_(course_ids)))
        await _del("offerings", delete(TimetableCourseOffering)
                   .where(TimetableCourseOffering.course_id.in_(course_ids)))
        await _del("entries", delete(TimetableEntry)
                   .where(TimetableEntry.course_id.in_(course_ids)))
        await _del("courses", delete(TimetableCourse)
                   .where(TimetableCourse.id.in_(course_ids)))
    elif entity == "rooms":
        await _del("entries", delete(TimetableEntry)
                   .where(TimetableEntry.room_id.in_(room_ids)))
        await _del("rooms", delete(TimetableRoom)
                   .where(TimetableRoom.id.in_(room_ids)))
    elif entity == "faculty":
        await _del("eligibility", delete(TimetableTeacherEligibility)
                   .where(TimetableTeacherEligibility.faculty_id.in_(faculty_ids)))
        await _del("entries", delete(TimetableEntry)
                   .where(TimetableEntry.faculty_id.in_(faculty_ids)))
        await _del("faculty", delete(TimetableFacultyProfile)
                   .where(TimetableFacultyProfile.id.in_(faculty_ids)))
    elif entity == "batches":
        await _del("entries", delete(TimetableEntry)
                   .where(TimetableEntry.section_id.in_(section_ids)))
        await _del("enrollments", delete(TimetableStudentEnrollment)
                   .where(TimetableStudentEnrollment.section_id.in_(section_ids)))
        await _del("lab_groups", delete(TimetableLabGroup)
                   .where(TimetableLabGroup.section_id.in_(section_ids)))
        await _del("offerings", delete(TimetableCourseOffering)
                   .where(TimetableCourseOffering.batch_id.in_(batch_ids)))
        await _del("sections", delete(TimetableSection)
                   .where(TimetableSection.batch_id.in_(batch_ids)))
        await _del("batches", delete(TimetableBatch)
                   .where(TimetableBatch.id.in_(batch_ids)))
    elif entity == "offerings":
        stmt = delete(TimetableCourseOffering)
        if term_id is not None:
            stmt = stmt.where(TimetableCourseOffering.term_id == term_id)
        elif tenant_id:
            stmt = stmt.where(TimetableCourseOffering.term_id.in_(term_ids))
        await _del("offerings", stmt)
    elif entity == "eligibility":
        stmt = delete(TimetableTeacherEligibility)
        if tenant_id:
            stmt = stmt.where(or_(
                TimetableTeacherEligibility.faculty_id.in_(faculty_ids),
                TimetableTeacherEligibility.course_id.in_(course_ids),
            ))
        await _del("eligibility", stmt)
    else:
        raise ValueError(f"unsupported entity '{entity}'")

    await db.commit()
    return deleted


def _parse_file(file_bytes: bytes, content_type: str) -> tuple[list[str], list[dict]]:
    """Parse a CSV/XLSX upload into ``(fieldnames, rows)``.

    Headers are stripped + lowercased so CSV and XLSX behave identically. Fully
    blank rows (common trailing rows in exported sheets) are dropped. CSV
    delimiter is auto-detected so files saved by Excel in non-US locales (which
    use ';') still import.
    """
    if "spreadsheet" in content_type or "xlsx" in content_type or "excel" in content_type:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            raw = list(ws.iter_rows(values_only=True))
            if not raw:
                return [], []
            headers = [str(h).strip().lower() if h is not None else "" for h in raw[0]]
            rows: list[dict] = []
            for r in raw[1:]:
                values = [(str(v).strip() if v is not None else "") for v in r]
                if not any(values):
                    continue  # skip fully-blank rows
                rows.append({
                    headers[j]: (values[j] if j < len(values) else "")
                    for j in range(len(headers))
                })
            return headers, rows
        except Exception as exc:
            raise ValueError(f"Failed to parse XLSX: {exc}")
    else:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text), delimiter=_sniff_delimiter(text))
        headers = [(h.strip().lower() if h else "") for h in (reader.fieldnames or [])]
        rows = []
        for raw_row in reader:
            row = {
                (k.strip().lower() if k else ""): (v.strip() if isinstance(v, str) else (v or ""))
                for k, v in raw_row.items()
            }
            if not any(str(v).strip() for v in row.values()):
                continue  # skip fully-blank rows
            rows.append(row)
        return headers, rows
